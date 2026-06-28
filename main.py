#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 26 20:05:26 2026

@author: a
"""
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
#from dotenv import load_dotenv
from pathlib import Path
from time import sleep
import xlsxwriter
import openpyxl
import os, sys, re, sqlite3

DATABASE_NAME = "sistema_usuarios.db"
usuarios = {}

def registrar_candidato(correo):
    """Registra un nuevo usuario en el sistema solicitando sus requerimientos mínimos."""
    print("\n--- REGISTRO DE CANDIDATO ---")
    print(f"Correo electrónico (Servirá como ID): {correo}")

    mail_regex_pattern = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,7}"
    if bool(re.findall(mail_regex_pattern, correo, re.IGNORECASE)):
        if correo in usuarios:
            print("Error: El usuario ya está registrado.")
        else:
            nombre = input("Nombre completo: ")
            idioma = input("Nivel de idioma dominante (Inglés/Español): ")
            salario = float(input("Expectativa salarial mínima mensual (Ej. 10000): "))
            password = input("Contraseña (Recomendado alfanumérica con caracteres especiales): ")
    
            usuarios[correo] = {
                "nombre": nombre,
                "idioma": idioma.capitalize(),
                "salario_esperado": salario,
                "password": password
            }
            
            # Crear base de datos sqlite
            with sqlite3.connect(DATABASE_NAME) as conn:
                cursor = conn.cursor()
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS usuarios (
                        correo TEXT PRIMARY KEY,
                        nombre TEXT NOT NULL,
                        idioma TEXT,
                        salario_esperado REAL,
                        password TEXT NOT NULL                
                    )
                """)
                
                query = """
                    INSERT OR REPLACE INTO usuarios (correo, nombre, idioma, salario_esperado, password)
                    VALUES (?, ?, ?, ?, ?)
                """
                
                for email, info in usuarios.items():
                    datos_registro = (
                        email,
                        info['nombre'], 
                        info['idioma'], 
                        info['salario_esperado'],
                        info['password']
                    )
                    cursor.execute(query, datos_registro)
                
                # Al salir del bloque 'with', se ejecuta conn.commit() automáticamente
                print("*** Usuario guardado exitosamente en la Base de Datos ***")
    else:
        print("NO se ingresó un correo válido")
    
    
def consultar_perfil(driver):
    """Permite al usuario ver sus datos registrados."""
    print("Logging success")        
    search_btn = driver.find_element(By.CSS_SELECTOR, '[data-testid="nav-search"]')
    search_btn.click()
    
    sleep(2)
    driver.find_element(By.CSS_SELECTOR, 'a[href="https://www.occ.com.mx/curriculo"]').click()
    
    sleep(2)
    nombre = driver.find_element(By.CSS_SELECTOR, "#personal-info p").text
    print(f"\nPerfil de: {nombre}")
 
    elementos_idiomas = driver.find_elements(By.CSS_SELECTOR, "#languages button span")
    lista_idiomas = [idioma.text for idioma in elementos_idiomas if idioma.text.strip()]
    print(f"Idioma: {' '.join(lista_idiomas)}")
    
    xpath_salario = "//div[@id='profession']//p[text()='Salario aproximado']/following-sibling::p"
    salario = driver.find_element(By.XPATH, xpath_salario).text
    print(f"Expectativa Salarial: {salario}")


def modificar_perfil(driver):
    """Cambiar la expectativa salarial minima"""
    nuevo_salario = float(input("Nueva expectativa salarial mínima: "))
    search_btn = driver.find_element(By.CSS_SELECTOR, '[data-testid="nav-search"]')
    search_btn.click()
    
    sleep(2)
    driver.find_element(By.CSS_SELECTOR, 'a[href="https://www.occ.com.mx/curriculo"]').click()
    
    xpath_lapiz = "//div[@id='profession']//*[contains(@class, 'atomic__pencil')]"
    elemento_lapiz = driver.find_element(By.XPATH, xpath_lapiz)

    actions = ActionChains(driver)
    actions.move_to_element(elemento_lapiz).click().perform()
    
    input_salario = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.ID, "salary")))
    
    input_salario.send_keys(Keys.CONTROL + "a")
    input_salario.send_keys(Keys.BACKSPACE)

    input_salario.send_keys(nuevo_salario)
    
    boton_guardar = driver.find_element(By.CSS_SELECTOR, "button[data-testid='saveData']")
    boton_guardar.click()
    print(f"Salario cambiado con éxito a: {nuevo_salario}")
    print("¡Perfil actualizado correctamente!")
        
    
def eliminar_perfil(correo):
    """Elimina el registro del usuario del sistema."""
    try:
        with sqlite3.connect(DATABASE_NAME) as conn:
            cursor = conn.cursor()
        
            cursor.execute("DELETE FROM usuarios WHERE correo = ?", (correo,))

        print(f"El usuario con correo {correo} ha sido eliminado.")
    except Exception as e:
        print(F"No se encontró ningún usuario con el correo {correo}.")


def excel_has_roles_to_search(excel_file_path:str):
    """Obtener datos filtrados"""
    # Search if exist excel input for to search roles in OCC 
    ls_roles_to_search = []
    
    wb = openpyxl.load_workbook(filename=excel_file_path)
    sh = wb["vacantes"]    
    
    for row in sh.iter_rows(values_only=True):
        if not row[0] == "Nombre de la vacante a buscar" and not row[0] is None:
            ls_roles_to_search.append(str(row[0]).strip())
    
    if bool(ls_roles_to_search):
        return ls_roles_to_search
    return []
        
        
def buscar_vacantes_match(excel_file_path:str, driver, usuario):
    """Cruza el perfil del usuario con las vacantes para mostrar solo las viables."""
    search_btn = driver.find_element(By.CSS_SELECTOR, '[data-testid="nav-search"]')
    search_btn.click()
    
    # Search if exist excel input for to search roles in OCC 
    ls_roles_to_search = []
    
    wb = openpyxl.load_workbook(filename=excel_file_path)
    sh = wb["vacantes"]    
    
    ls_roles_to_search = excel_has_roles_to_search(excel_file_path=excel_file_path)
    if not bool(ls_roles_to_search):
        raise Exception(f"Especificar las vacantes que desea buscar por nombre o palabra clave en archivo '{excel_file_path}'")

    # Search Roles position in OCC search input
    for role in ls_roles_to_search:
        search_role_input = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "prof-cat-search-input-desktop")))
        search_role_input.send_keys(Keys.CONTROL, "a")
        search_role_input.send_keys(Keys.BACKSPACE)

        search_role_input.send_keys(role.lower())
        search_role_input.send_keys(Keys.ENTER) 
        sleep(4)
        
        all_jobs_search = driver.find_elements(By.CSS_SELECTOR, "div[data-offers-grid-offer-item-container]")

        if bool(all_jobs_search):
            print("********** EMPLEOS ENCONTRADOS **********")
            for job_pos in all_jobs_search:
                try:
                    cover_salary = False
                    cover_language = False
                    
                    # Get info job position
                    job_name = job_pos.find_element(By.TAG_NAME, "h2").text.strip()
                    print(f" - {job_name}")
                    
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", job_pos)
                    sleep(1)

                    driver.execute_script("arguments[0].click();", job_pos)
                    sleep(2)

                    job_pos_text = driver.find_element(By.ID, "job-detail-container").text

                    try:
                        # Get salary
                        salary = re.findall(r"(\d[\d,.]*)\s*(?:-|–|—|[aA]\b|al\b|hasta\b)\s*(?:[^0-9\n]*)\s*(\d[\d,.]*)", job_pos)
                        min_salary = salary[0][0]
                        max_salary = salary[0][1]
                        
                        if ',' in min_salary:    min_salary = float(min_salary.replace(',', ''))
                        if ',' in max_salary:    max_salary = float(max_salary.replace(',', ''))                        
                            
                        print(f"- Salario minimo {min_salary}\n- Salario maximo {max_salary}")
                        
                        if float(min_salary) >= usuario["salario_esperado"]:
                            cover_salary = True
                    except:
                        print("Salario no mostrado")
                    
                    # Get languaje
                    if usuario["idioma"].lower() == "ingles":
                        english_pos = re.findall(r"(ingles|english|working in the United States)", job_pos, re.IGNORECASE)
                        if english_pos:
                            cover_language = True

                    else:
                        cover_language = True
                        
                    if cover_language and cover_salary:
                        print("POSTULARSE")
                        
                    else:
                        print("NO POSTULARSE: no se cumplen los requisitos")
                    
                except StaleElementReferenceException:
                    continue
            print("===============================================")
    
    if True:
        print("Por el momento no hay vacantes que cumplan con tus requisitos exactos.")
    else:
        print("Debes estar registrado para buscar vacantes.")


def logging_page(driver:webdriver, user:str, password:str):
    """Ingresar a OCC mundial"""
    try:
        page_load = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "login-link")))        
        if bool(page_load.text):
            page_load.click()

            # Insert credential to web page            
            user_cred = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "inputID_identifier")))
            user_cred.send_keys(user)
            
            pass_cred = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "inputID_password")))
            pass_cred.send_keys(password)
            sleep(2)

            btn_logging = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@id='inputID_method']")))
            btn_logging.click()

            main_user_cv = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "cv-views")))
                        
            if bool(main_user_cv.text):
                return True
        
        return False

    except Exception as e:
        print(f"Error: {e}")
        return False

        
def open_web_page(driver:webdriver, url_web_page:str):
    """Abrir Navegador"""
    driver.get(url_web_page)


def excel_search_role_pos(excel_file_path:str):
    """Crear excel para validar puestoa a aplicar"""
    try:
        # Validate if excel file exists
        if not os.path.isfile(excel_file_path):
            wb = xlsxwriter.Workbook(excel_file_path)
            ws = wb.add_worksheet("vacantes")
            ws.write(0, 0, 'Nombre de la vacante a buscar')
            wb.close()
            return False
        return True
    
    except Exception as e:
        print(f"Error: {e}")
        return False


def obtener_datos_usuario(correo):
    """Busca un usuario por su correo en la base de datos."""
    try:
        with sqlite3.connect(DATABASE_NAME) as conn:
            cursor = conn.cursor()
            
            # Traemos las 5 columnas correspondientes de la tabla
            query = """
                SELECT correo, nombre, idioma, salario_esperado, password 
                FROM usuarios 
                WHERE correo = ?
            """
            cursor.execute(query, (correo,))
            resultado = cursor.fetchone()
            
            if resultado:
                datos_usuario = {
                    "correo": resultado[0],
                    "nombre": resultado[1],
                    "idioma": resultado[2],
                    "salario_esperado": resultado[3],
                    "password": resultado[4]
                }
                return datos_usuario
            else:
                return None
                    
    except sqlite3.OperationalError:
        # Por si la tabla aún no ha sido creada en la base de datos
        print("Error: La tabla 'usuarios' no existe todavía.")
        return None

def menu_principal():
    """Estructura de control principal del programa.""" 
    # Load variables
    url_occ_mundial  = "https://www.occ.com.mx/"
    excel_file       = "busqueda_vacantes.xlsx"

    # Ruta origen
    base_path = str(Path(__file__).parent)
    excel_file_path = os.path.join(base_path, excel_file)

    opcion = ""
    while opcion != "5":
        print("\n=== SMART MATCHING CCC ===")
        print("1. Consultar Perfil")
        print("2. Modificar Perfil")
        print("3. Eliminar Perfil")
        print("4. Buscar Vacantes (Smart Match)")
        print("5. Salir")

        opcion = input("Selecciona una opción: ")
        
        if bool(re.findall("[1-5]", opcion)):
            if int(opcion) == 5:    break
                        
            mail_access = input("Ingrese su correo para ingresar a cuenta: ")
            
            # Establecer conexión 
            with sqlite3.connect(DATABASE_NAME) as conn:
                cursor = conn.cursor()
        
                cursor.execute("SELECT count(*) FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
                total_tablas = cursor.fetchone()[0]

            if total_tablas == 0:
                print("La base de datos está vacía (sin tablas).")
            else:
                # Retry empty user
                while True:
                    usuario = obtener_datos_usuario(correo=mail_access)                    
                    if bool(usuario):
                        break
                    registrar_candidato(correo=mail_access)
                
            mail_regex_pattern = "[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,7}"
            if bool(re.findall(mail_regex_pattern, mail_access, re.IGNORECASE)):
                # Create the instance of driver (Who control the web page)
                service = Service(ChromeDriverManager().install())
                driver  = webdriver.Chrome(service=service)
                    
                # Wait for o load the web page
                driver.implicitly_wait(30)
                driver.maximize_window()
                
                # Enter to the web page
                open_web_page(driver=driver, url_web_page=url_occ_mundial)            
                logging_result = logging_page(driver=driver, user=usuario["correo"], password=usuario["password"])
                
                if logging_result:
                    if opcion == "1":
                        consultar_perfil(driver)
                    elif opcion == "2":
                        modificar_perfil(driver)
                    elif opcion == "3":
                        eliminar_perfil(mail_access)
                    elif opcion == "4":
                        buscar_vacantes_match(excel_file_path, driver, usuario)
                    elif opcion == "5":
                        print("Saliendo del sistema...")
                else:
                    print("No se pudo ingresar a OCC")
                driver.close()
                driver.quit()
                
        else:
            print("Opción no válida. Intenta de nuevo.")


# Ejecución del programa
if __name__ == "__main__":
    menu_principal()